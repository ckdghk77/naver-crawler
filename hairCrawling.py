'''
Created on Jun 13, 2016

@author: TYchoi
'''

from selenium import webdriver
from bs4 import BeautifulSoup
import xlsxwriter
from urllib import parse
from openpyxl.compat import range
from openpyxl import load_workbook

def writeToExcel(wantToSave,name):
    numberOfRows = len(wantToSave)
    numberOfCols = len(wantToSave[0])
    workbook = xlsxwriter.Workbook(name+'.xlsx')
    workseet = workbook.add_worksheet()
    for i in range(numberOfRows):
        for j in range(numberOfCols):
            workseet.write(i,j, str(wantToSave[i][j]))

def maxNumberFinder(soup):
    numOfPosts = soup.find('span',{'class':'title_num'})
    
    if numOfPosts:
        numOfPosts = int(numOfPosts.text.split(' ')[-1].replace(',','')[:-1])   
        if numOfPosts > 50:
            maxnum = 5
        else:
            if numOfPosts%10 == 0:
                maxnum = int(numOfPosts/10)
            else:
                maxnum = int(numOfPosts/10)+1
    else:
        numOfPosts = 0
        maxnum = 0   
    return numOfPosts, maxnum

def urlGenerator(hairStyle, fromDate, toDate, pageNum):
    query=hairStyle.replace(hairStyle,parse.quote(hairStyle))
    searchUrl='https://search.naver.com/search.naver?where=post&query='+query+'&ie=utf8&st=sim&sm=tab_opt&date_from='+fromDate+'&date_to='+toDate+'&date_option=6&srchby=all&dup_remove=1&post_blogurl=&post_blogurl_without=&nso=so%3Ar%2Ca%3Aall%2Cp%3Afrom'+fromDate+'to'+toDate+'&mson=0'
    searchUrl = searchUrl + '&start=' + pageNum
    return searchUrl

def hairStyleReader():
    wb = load_workbook(filename = '헤어.xlsx')
    sheet_ranges =  wb['Sheet1']
    hairStyleList = []
    for i in range (4, 103):
        hairStyleList.append([sheet_ranges['A'+str(i)].value,sheet_ranges['B'+str(i)].value,sheet_ranges['C'+str(i)].value])
    return hairStyleList

def dateRageGenerator():
    dateRage = [['20130601', '20130831'],['20130901', '20131130'],['20131201', '20140228'],['20140301', '20140531'],
                ['20140601', '20140831'],['20140901', '20141130'],['20141201', '20150228'],['20150301', '20150531'],
                ['20150601', '20150831'],['20150901', '20151130'],['20151201', '20160228'],['20160301', '20160531']]    
    return dateRage

def pageSourceRetriever(driver, searchUrl):
    driver.get(searchUrl)
    r=driver.page_source
    soup=BeautifulSoup(r, "lxml") 
    return soup

def getHrefs(soup):
    for item in soup.find_all('a', {'class':'sh_blog_title _sp_each_url _sp_each_title'}):
        href=item['href']
    return href

def getDate(soup):
    for item in soup.find_all('dd',{'class':'txt_inline'}):
        date = item.text.split(" ")[0]
    return date

def main():
    driver = webdriver.PhantomJS(executable_path='/Users/TYchoi/PythonProjects/phantomjs/bin/phantomjs')
    hairStyleList = hairStyleReader()
    dateRange = dateRageGenerator()
    overallInfo = []
    listOfPosts = []
    for hairStyle in hairStyleList:
        print (hairStyle[1])
        for date in dateRange:
            searchUrl = urlGenerator(hairStyle[2], date[0], date[1], '')
            soup = pageSourceRetriever(driver, searchUrl) 
            numOfPosts, maxNum = maxNumberFinder(soup)
            
            for i in range(maxNum):
                num = str(10*i+1)
                searchUrlByPage = urlGenerator(hairStyle[2], date[0], date[1], num)
                soup = pageSourceRetriever(driver, searchUrlByPage)
                
                sections = soup.find_all('li', {'class':'sh_blog_top'})
                    
                for section in sections:
                    href = getHrefs(section)
                    datePosted = getDate(section)
                    listOfPosts.append([hairStyle[0],hairStyle[1], datePosted, date[0] ,href])
            overallInfo.append([hairStyle[0],hairStyle[1],date[0], numOfPosts])
    writeToExcel(listOfPosts, '헤어블로그리스트')
    writeToExcel(overallInfo, '버즈량분석')
main()